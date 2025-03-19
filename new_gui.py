import customtkinter as ctk
import logging
import os
import threading
import time
import json
import subprocess
from concurrent.futures import ThreadPoolExecutor
from parsers.site1_parser import Site1Parser
from parsers.site2_parser import Site2Parser
from parsers.site3_parser import Site3Parser
from tkinter import messagebox, filedialog
from tkinter import ttk
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from review_analyzer import ReviewAnalyzer
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import Counter
from wordcloud import WordCloud

# Константы
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
CONFIG_FILE = "parser_config.json"
WINDOW_SIZE = "900x700"
LOG_HEIGHT = 12
SITE_NAMES = {
    "site1": "Отзовик",
    "site2": "iRecommend",
    "site3": "Озон"
}

class InfoFilter(logging.Filter):
    def filter(self, record):
        return record.levelno == logging.INFO

class ParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Парсер отзывов")
        self.root.geometry(WINDOW_SIZE)
        self.root.resizable(True, True)
        self.root.configure(bg="#1A1A1A")
        self.running = False  # Флаг для парсинга
        self.is_running = True  # Глобальный флаг для отслеживания работы приложения
        self.threads = []  # Список для хранения всех потоков

        self.output_dir = ctk.StringVar(value=DATA_DIR)
        self.site_links = {site: ctk.StringVar() for site in SITE_NAMES}
        self.use_preprocessing = ctk.BooleanVar(value=True)
        self.csv_files = []
        self.product_name = ctk.StringVar(value="продукт")
        self.analyzer = ReviewAnalyzer(use_preprocessing=self.use_preprocessing.get())
        self.results_by_site = []
        self.aggregated_results = None
        self.analysis_cache = {}  # Для кэширования результатов анализа

        self.create_widgets()
        self.setup_logger_redirect()
        self.load_config()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        ctk.set_appearance_mode("dark")
        print(f"Текущая тема: {ctk.get_appearance_mode()}")
        self.bind_shortcuts()

    def create_widgets(self):
        main_frame = ctk.CTkFrame(self.root, fg_color="#1A1A1A")
        main_frame.pack(padx=5, pady=5, fill="both", expand=True)

        self.tabview = ctk.CTkTabview(main_frame, width=880, height=650, fg_color="#2C2C2C", bg_color="#1A1A1A", border_color="#1DA1F2")
        self.tabview.pack(padx=5, pady=5, fill="both", expand=True)
        self.tabview.add("Парсер отзывов")
        self.tabview.add("Анализ слов")
        self.tabview.add("Детальный анализ предложений")
        self.tabview.add("Текстовый отчёт")

        # Вкладка "Парсер отзывов"
        parser_tab = self.tabview.tab("Парсер отзывов")
        
        links_frame = ctk.CTkFrame(parser_tab, fg_color="#2C2C2C")
        links_frame.pack(fill="x", pady=2, padx=5, expand=False)
        ctk.CTkLabel(links_frame, text="Ссылки на отзывы", font=("Arial", 14, "bold"), text_color="white").pack(anchor="w", padx=5, pady=2)
        
        for site in SITE_NAMES:
            site_frame = ctk.CTkFrame(links_frame, fg_color="#2C2C2C")
            site_frame.pack(fill="x", pady=2, padx=5)
            site_frame.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(site_frame, text=SITE_NAMES[site], font=("Arial", 12), text_color="white").grid(row=0, column=0, padx=5, pady=5, sticky="w")
            ctk.CTkEntry(site_frame, textvariable=self.site_links[site], width=600, height=30, font=("Arial", 12), fg_color="#3A3A3A", border_color="#1DA1F2", corner_radius=10).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
            ctk.CTkButton(site_frame, text="Добавить", command=lambda s=site: self.add_link(s), fg_color="#1DA1F2", hover_color="#166AB1", width=80, height=30, font=("Arial", 12), corner_radius=10).grid(row=0, column=2, padx=5, pady=5, sticky="e")

        output_frame = ctk.CTkFrame(parser_tab, fg_color="#2C2C2C")
        output_frame.pack(fill="x", pady=2, padx=5, expand=False)
        ctk.CTkLabel(output_frame, text="Папка сохранения", font=("Arial", 14, "bold"), text_color="white").pack(anchor="w", padx=5, pady=2)
        inner_frame = ctk.CTkFrame(output_frame, fg_color="#2C2C2C")
        inner_frame.pack(fill="x", padx=5, pady=2)
        ctk.CTkEntry(inner_frame, textvariable=self.output_dir, width=600, height=30, font=("Arial", 12), fg_color="#3A3A3A", border_color="#1DA1F2", corner_radius=10).pack(side="left", fill="x", expand=True, padx=5)
        ctk.CTkButton(inner_frame, text="Обзор...", command=self.select_output_dir, fg_color="#1DA1F2", hover_color="#166AB1", width=80, height=30, font=("Arial", 12), corner_radius=10).pack(side="right", padx=5)

        log_frame = ctk.CTkFrame(parser_tab, fg_color="#2C2C2C")
        log_frame.pack(fill="both", expand=True, pady=2, padx=5)
        ctk.CTkLabel(log_frame, text="Лог выполнения", font=("Arial", 14, "bold"), text_color="white").pack(anchor="w", padx=5, pady=2)
        self.log_text = ctk.CTkTextbox(log_frame, height=LOG_HEIGHT, width=850, state="disabled", font=("Arial", 12), text_color="white", fg_color="#3A3A3A", border_color="#1DA1F2")
        scrollbar = ctk.CTkScrollbar(log_frame, command=self.log_text.yview, fg_color="#2C2C2C", button_color="#1DA1F2")
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side="left", fill="both", expand=True, padx=5)
        scrollbar.pack(side="right", fill="y")

        control_frame = ctk.CTkFrame(parser_tab, fg_color="#1A1A1A")
        control_frame.pack(fill="x", pady=5, padx=5, expand=False)
        self.start_btn = ctk.CTkButton(control_frame, text="Старт", command=self.toggle_parsing, fg_color="#1DA1F2", hover_color="#166AB1", width=80, height=30, font=("Arial", 12), corner_radius=10)
        self.start_btn.pack(side="left", padx=5)
        ctk.CTkButton(control_frame, text="Очистить лог", command=self.clear_log, fg_color="#4A4A4A", hover_color="#5A5A5A", width=80, height=30, font=("Arial", 12), corner_radius=10).pack(side="left", padx=5)
        ctk.CTkButton(control_frame, text="Открыть папку", command=self.open_output_dir, fg_color="#4A4A4A", hover_color="#5A5A5A", width=80, height=30, font=("Arial", 12), corner_radius=10).pack(side="left", padx=5)
        self.progress = ctk.CTkProgressBar(control_frame, mode="determinate", width=250, height=20, progress_color="#1DA1F2", fg_color="#4A4A4A")
        self.progress.pack(side="right", padx=5)
        self.progress.set(0)

        # Вкладка "Анализ слов"
        analysis_tab = self.tabview.tab("Анализ слов")

        file_frame = ctk.CTkFrame(analysis_tab, fg_color="#2C2C2C")
        file_frame.pack(fill="x", pady=2, padx=5, expand=False)
        ctk.CTkLabel(file_frame, text="Выбор CSV-файлов для анализа", font=("Arial", 14, "bold"), text_color="white").pack(anchor="w", padx=5, pady=2)
        ctk.CTkCheckBox(file_frame, text="Использовать предобработку текста (исправление опечаток)", variable=self.use_preprocessing, font=("Arial", 12), text_color="white", command=self.update_analyzer).pack(anchor="w", padx=5, pady=2)
        ctk.CTkButton(file_frame, text="Выбрать файлы для анализа по сайтам", command=self.select_files_for_analysis_by_site, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(pady=5)
        ctk.CTkButton(file_frame, text="Агрегированный анализ выбранных файлов", command=self.select_files_for_aggregated_analysis, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(pady=5)
        ctk.CTkButton(file_frame, text="Детальный анализ предложений", command=self.select_files_for_detailed_analysis, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(pady=5)
        ctk.CTkButton(file_frame, text="Сохранить таблицу в Excel", command=self.save_table_to_excel, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(pady=5)
        ctk.CTkButton(file_frame, text="Визуализировать гистограмму", command=self.visualize_results, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(pady=5)
        ctk.CTkButton(file_frame, text="Визуализировать облако слов", command=self.visualize_wordcloud, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(pady=5)

        result_frame = ctk.CTkFrame(analysis_tab, fg_color="#2C2C2C")
        result_frame.pack(fill="both", expand=True, pady=5, padx=5)
        ctk.CTkLabel(result_frame, text="Результаты анализа", font=("Arial", 14, "bold"), text_color="white").pack(anchor="w", padx=5, pady=2)

        tree_frame = ctk.CTkFrame(result_frame, fg_color="#333333")
        tree_frame.pack(fill="both", expand=True)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                        background="#333333",
                        foreground="#FFFFFF",
                        fieldbackground="#333333",
                        rowheight=150,
                        font=("Arial", 12))
        style.configure("Custom.Treeview.Heading",
                        background="#444444",
                        foreground="#FFFFFF",
                        font=("Arial", 12, "bold"))
        style.map("Custom.Treeview",
                  background=[("selected", "#1DA1F2")],
                  foreground=[("selected", "#FFFFFF")])

        self.result_tree = ttk.Treeview(tree_frame, columns=("Category", "Pros", "Cons", "PositiveKeywords", "NegativeKeywords", "CommonKeywords", "PositiveReviews", "NegativeReviews"), show="headings", style="Custom.Treeview")
        self.result_tree.heading("Category", text="Категория")
        self.result_tree.heading("Pros", text="Плюсы")
        self.result_tree.heading("Cons", text="Минусы")
        self.result_tree.heading("PositiveKeywords", text="Ключевые слова (положительные)")
        self.result_tree.heading("NegativeKeywords", text="Ключевые слова (отрицательные)")
        self.result_tree.heading("CommonKeywords", text="Общие ключевые слова")
        self.result_tree.heading("PositiveReviews", text="Положительные отзывы")
        self.result_tree.heading("NegativeReviews", text="Отрицательные отзывы")

        self.result_tree.column("Category", width=150, anchor="w")
        self.result_tree.column("Pros", width=250, anchor="w")
        self.result_tree.column("Cons", width=250, anchor="w")
        self.result_tree.column("PositiveKeywords", width=300, anchor="w")
        self.result_tree.column("NegativeKeywords", width=300, anchor="w")
        self.result_tree.column("CommonKeywords", width=300, anchor="w")
        self.result_tree.column("PositiveReviews", width=100, anchor="w")
        self.result_tree.column("NegativeReviews", width=100, anchor="w")

        style.configure("Custom.Vertical.TScrollbar",
                        troughcolor="#1E1E1E",
                        background="#1DA1F2",
                        arrowcolor="#FFFFFF")
        v_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', style="Custom.Vertical.TScrollbar", command=self.result_tree.yview)
        v_scrollbar.pack(side="right", fill="y")
        self.result_tree.configure(yscrollcommand=v_scrollbar.set)

        style.configure("Custom.Horizontal.TScrollbar",
                        troughcolor="#1E1E1E",
                        background="#1DA1F2",
                        arrowcolor="#FFFFFF")
        h_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', style="Custom.Horizontal.TScrollbar", command=self.result_tree.xview)
        h_scrollbar.pack(side="bottom", fill="x")
        self.result_tree.configure(xscrollcommand=h_scrollbar.set)

        self.result_tree.pack(fill="both", expand=True)

        # Вкладка "Детальный анализ предложений"
        detailed_tab = self.tabview.tab("Детальный анализ предложений")

        detailed_frame = ctk.CTkFrame(detailed_tab, fg_color="#2C2C2C")
        detailed_frame.pack(fill="both", expand=True, pady=5, padx=5)
        ctk.CTkLabel(detailed_frame, text="Детальный анализ предложений", font=("Arial", 14, "bold"), text_color="white").pack(anchor="w", padx=5, pady=2)

        button_frame = ctk.CTkFrame(detailed_frame, fg_color="#2C2C2C")
        button_frame.pack(fill="x", pady=5, padx=5)
        ctk.CTkButton(button_frame, text="Сохранить детальный анализ в Excel", command=self.save_detailed_table_to_excel, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(side="left", padx=5)
        self.progress_detailed = ctk.CTkProgressBar(button_frame, mode="determinate", width=250, height=20, progress_color="#1DA1F2", fg_color="#4A4A4A")
        self.progress_detailed.pack(side="right", padx=5)
        self.progress_detailed.set(0)

        tree_frame = ctk.CTkFrame(detailed_frame, fg_color="#333333")
        tree_frame.pack(fill="both", expand=True)

        self.detailed_tree = ttk.Treeview(tree_frame, columns=("Sentence", "Sentiment", "Score", "Aspects"), show="headings", style="Custom.Treeview")
        self.detailed_tree.heading("Sentence", text="Предложение")
        self.detailed_tree.heading("Sentiment", text="Тональность")
        self.detailed_tree.heading("Score", text="Скор")
        self.detailed_tree.heading("Aspects", text="Аспекты")

        self.detailed_tree.column("Sentence", width=300, anchor="w")
        self.detailed_tree.column("Sentiment", width=100, anchor="w")
        self.detailed_tree.column("Score", width=100, anchor="w")
        self.detailed_tree.column("Aspects", width=400, anchor="w")

        v_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', style="Custom.Vertical.TScrollbar", command=self.detailed_tree.yview)
        v_scrollbar.pack(side="right", fill="y")
        self.detailed_tree.configure(yscrollcommand=v_scrollbar.set)

        h_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', style="Custom.Horizontal.TScrollbar", command=self.detailed_tree.xview)
        h_scrollbar.pack(side="bottom", fill="x")
        self.detailed_tree.configure(xscrollcommand=h_scrollbar.set)

        self.detailed_tree.pack(fill="both", expand=True)

        # Вкладка "Текстовый отчёт"
        report_tab = self.tabview.tab("Текстовый отчёт")

        report_frame = ctk.CTkFrame(report_tab, fg_color="#2C2C2C")
        report_frame.pack(fill="both", expand=True, pady=5, padx=5)
        ctk.CTkLabel(report_frame, text="Текстовый отчёт", font=("Arial", 14, "bold"), text_color="white").pack(anchor="w", padx=5, pady=2)

        product_frame = ctk.CTkFrame(report_frame, fg_color="#2C2C2C")
        product_frame.pack(fill="x", pady=2, padx=5)
        ctk.CTkLabel(product_frame, text="Название продукта:", font=("Arial", 12), text_color="white").pack(side="left", padx=5)
        ctk.CTkEntry(product_frame, textvariable=self.product_name, width=300, height=30, font=("Arial", 12), fg_color="#3A3A3A", border_color="#1DA1F2", corner_radius=10).pack(side="left", padx=5)

        button_frame = ctk.CTkFrame(report_frame, fg_color="#2C2C2C")
        button_frame.pack(fill="x", pady=5, padx=5)
        ctk.CTkButton(button_frame, text="Сгенерировать отчёт по сайтам", command=self.generate_report_by_site, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="Сгенерировать агрегированный отчёт", command=self.generate_aggregated_report, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="Сохранить отчёт в файл", command=self.save_report_to_file, fg_color="#1DA1F2", hover_color="#166AB1", width=250, height=30, font=("Arial", 12), corner_radius=10).pack(side="left", padx=5)
        self.progress_report = ctk.CTkProgressBar(button_frame, mode="determinate", width=250, height=20, progress_color="#1DA1F2", fg_color="#4A4A4A")
        self.progress_report.pack(side="right", padx=5)
        self.progress_report.set(0)

        self.report_textbox = ctk.CTkTextbox(report_frame, height=400, width=850, font=("Arial", 12), text_color="white", fg_color="#3A3A3A", border_color="#1DA1F2")
        self.report_textbox.pack(fill="both", expand=True, padx=5, pady=5)

    def update_analyzer(self):
        if not self.is_running:
            return
        self.analyzer = ReviewAnalyzer(use_preprocessing=self.use_preprocessing.get())
        self.log(f"Анализатор обновлён с предобработкой: {self.use_preprocessing.get()}")

    def save_table_to_excel(self):
        if not self.is_running:
            return
        columns = ["Категория", "Плюсы", "Минусы", "Ключевые слова (положительные)", "Ключевые слова (отрицательные)", "Общие ключевые слова", "Положительные отзывы", "Отрицательные отзывы"]
        data = []

        for item in self.result_tree.get_children():
            values = self.result_tree.item(item, "values")
            data.append(values)

        if not data:
            self.log("Таблица пуста, нечего сохранять!", is_error=True)
            messagebox.showwarning("Предупреждение", "Таблица пуста, нечего сохранять!")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить таблицу как"
        )

        if not file_path:
            self.log("Сохранение отменено пользователем.")
            return

        try:
            df = pd.DataFrame(data, columns=columns)
            df.to_excel(file_path, index=False, engine='openpyxl')
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            for col in range(1, len(columns) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{column_letter}{row}"]
                    try:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell_value = str(cell.value)
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
                adjusted_width = max_length + 5
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            self.log(f"Таблица успешно сохранена в {file_path}")
            messagebox.showinfo("Успех", f"Таблица сохранена в {file_path}")
        except Exception as e:
            self.log(f"Ошибка при сохранении таблицы в Excel: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось сохранить таблицу: {str(e)}")

    def save_detailed_table_to_excel(self):
        if not self.is_running:
            return
        columns = ["Предложение", "Тональность", "Скор", "Аспекты"]
        data = []

        for item in self.detailed_tree.get_children():
            values = self.detailed_tree.item(item, "values")
            data.append(values)

        if not data:
            self.log("Таблица детального анализа пуста, нечего сохранять!", is_error=True)
            messagebox.showwarning("Предупреждение", "Таблица детального анализа пуста, нечего сохранять!")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить детальный анализ как"
        )

        if not file_path:
            self.log("Сохранение отменено пользователем.")
            return

        try:
            df = pd.DataFrame(data, columns=columns)
            df.to_excel(file_path, index=False, engine='openpyxl')
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            for col in range(1, len(columns) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{column_letter}{row}"]
                    try:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell_value = str(cell.value)
                        max_length = max(max_length, len(cell_value))
                    except:
                        pass
                adjusted_width = max_length + 5
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            self.log(f"Детальный анализ успешно сохранён в {file_path}")
            messagebox.showinfo("Успех", f"Детальный анализ сохранён в {file_path}")
        except Exception as e:
            self.log(f"Ошибка при сохранении детального анализа в Excel: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось сохранить детальный анализ: {str(e)}")

    def save_report_to_file(self):
        if not self.is_running:
            return
        report_content = self.report_textbox.get("0.0", "end").strip()
        if not report_content:
            self.log("Отчёт пуст, нечего сохранять!", is_error=True)
            messagebox.showwarning("Предупреждение", "Отчёт пуст, нечего сохранять!")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            title="Сохранить отчёт как"
        )

        if not file_path:
            self.log("Сохранение отменено пользователем.")
            return

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(report_content)
            self.log(f"Отчёт успешно сохранён в {file_path}")
            messagebox.showinfo("Успех", f"Отчёт сохранён в {file_path}")
        except Exception as e:
            self.log(f"Ошибка при сохранении отчёта: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось сохранить отчёт: {str(e)}")

    def visualize_results(self):
        if not self.is_running:
            return
        if not self.csv_files:
            self.log("Не выбраны CSV-файлы для анализа, нечего визуализировать!", is_error=True)
            messagebox.showwarning("Предупреждение", "Не выбраны CSV-файлы для анализа!")
            return

        self.progress.set(0)
        self.log("Запуск визуализации гистограммы...")

        def compute_data():
            try:
                all_reviews = []
                for i, csv_file in enumerate(self.csv_files):
                    if not self.is_running:
                        return None, None
                    try:
                        df = pd.read_csv(csv_file, encoding='utf-8')
                        if 'Текст отзыва' in df.columns:
                            reviews = df['Текст отзыва'].dropna().tolist()
                            all_reviews.extend(reviews)
                        elif 'Достоинства' in df.columns and 'Недостатки' in df.columns:
                            pros_reviews = df['Достоинства'].dropna().tolist()
                            cons_reviews = df['Недостатки'].dropna().tolist()
                            all_reviews.extend(pros_reviews + cons_reviews)
                        else:
                            self.log(f"CSV-файл {csv_file} не содержит подходящих столбцов для анализа!", is_error=True)
                            continue
                    except Exception as e:
                        self.log(f"Ошибка при чтении файла {csv_file}: {str(e)}", is_error=True)
                        continue
                    self.progress.set((i + 1) / len(self.csv_files) * 100)

                if not all_reviews:
                    self.log("Нет отзывов для визуализации!", is_error=True)
                    messagebox.showwarning("Предупреждение", "Нет отзывов для визуализации!")
                    return None, None

                positive_keywords = []
                negative_keywords = []
                keyword_review_count = Counter()

                for review in all_reviews:
                    if not self.is_running:
                        return None, None
                    sentences = self.analyzer.analyze_review_sentences(review)
                    review_pos_keywords = set()
                    review_neg_keywords = set()
                    for _, _, _, aspects in sentences:
                        for aspect_phrase, aspect_sentiment, _, _ in aspects:
                            if aspect_phrase.lower() in self.analyzer.invalid_phrases:
                                continue
                            if aspect_sentiment == "положительное":
                                positive_keywords.append(aspect_phrase)
                                review_pos_keywords.add(aspect_phrase)
                            elif aspect_sentiment == "отрицательное":
                                negative_keywords.append(aspect_phrase)
                                review_neg_keywords.add(aspect_phrase)
                    for keyword in review_pos_keywords:
                        keyword_review_count[keyword] += 1
                    for keyword in review_neg_keywords:
                        keyword_review_count[keyword] += 1

                pos_keyword_counts = Counter(positive_keywords)
                neg_keyword_counts = Counter(negative_keywords)

                MIN_REVIEW_THRESHOLD = 2
                filtered_pos_keywords = {k: v for k, v in pos_keyword_counts.items() if keyword_review_count[k] >= MIN_REVIEW_THRESHOLD}
                filtered_neg_keywords = {k: v for k, v in neg_keyword_counts.items() if keyword_review_count[k] >= MIN_REVIEW_THRESHOLD}

                return filtered_pos_keywords, filtered_neg_keywords
            except Exception as e:
                self.log(f"Ошибка при вычислении данных для визуализации: {str(e)}", is_error=True)
                messagebox.showerror("Ошибка", f"Не удалось выполнить вычисления: {str(e)}")
                return None, None
            finally:
                self.progress.set(0)

        def show_visualization(filtered_pos_keywords, filtered_neg_keywords):
            if not self.is_running:
                return

            if filtered_pos_keywords is None or filtered_neg_keywords is None:
                return

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))

            if filtered_pos_keywords:
                bars = ax1.barh(list(filtered_pos_keywords.keys())[:10], list(filtered_pos_keywords.values())[:10], color='green')
                ax1.set_title("Топ-10 положительных ключевых слов")
                ax1.set_xlabel("Частота")
                ax1.invert_yaxis()
                for bar in bars:
                    width = bar.get_width()
                    ax1.text(width, bar.get_y() + bar.get_height()/2, f'{int(width)}', ha='left', va='center')

            if filtered_neg_keywords:
                bars = ax2.barh(list(filtered_neg_keywords.keys())[:10], list(filtered_neg_keywords.values())[:10], color='red')
                ax2.set_title("Топ-10 отрицательных ключевых слов")
                ax2.set_xlabel("Частота")
                ax2.invert_yaxis()
                for bar in bars:
                    width = bar.get_width()
                    ax2.text(width, bar.get_y() + bar.get_height()/2, f'{int(width)}', ha='left', va='center')

            plt.tight_layout()

            top = ctk.CTkToplevel(self.root)
            top.title("Гистограмма ключевых слов")
            canvas = FigureCanvasTkAgg(fig, master=top)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)

            self.log("Визуализация гистограммы завершена!")

        def run():
            filtered_pos, filtered_neg = compute_data()
            if filtered_pos is not None and filtered_neg is not None and self.is_running:
                self.root.after(0, show_visualization, filtered_pos, filtered_neg)

        thread = threading.Thread(target=run)
        self.threads.append(thread)
        thread.start()

    def visualize_wordcloud(self):
        if not self.is_running:
            return
        if not self.csv_files:
            self.log("Не выбраны CSV-файлы для анализа, нечего визуализировать!", is_error=True)
            messagebox.showwarning("Предупреждение", "Не выбраны CSV-файлы для анализа!")
            return

        self.progress.set(0)
        self.log("Запуск визуализации облака слов...")

        def compute_data():
            try:
                all_reviews = []
                for i, csv_file in enumerate(self.csv_files):
                    if not self.is_running:
                        return None, None
                    try:
                        df = pd.read_csv(csv_file, encoding='utf-8')
                        if 'Текст отзыва' in df.columns:
                            reviews = df['Текст отзыва'].dropna().tolist()
                            all_reviews.extend(reviews)
                        elif 'Достоинства' in df.columns and 'Недостатки' in df.columns:
                            pros_reviews = df['Достоинства'].dropna().tolist()
                            cons_reviews = df['Недостатки'].dropna().tolist()
                            all_reviews.extend(pros_reviews + cons_reviews)
                        else:
                            self.log(f"CSV-файл {csv_file} не содержит подходящих столбцов для анализа!", is_error=True)
                            continue
                    except Exception as e:
                        self.log(f"Ошибка при чтении файла {csv_file}: {str(e)}", is_error=True)
                        continue
                    self.progress.set((i + 1) / len(self.csv_files) * 100)

                if not all_reviews:
                    self.log("Нет отзывов для визуализации!", is_error=True)
                    messagebox.showwarning("Предупреждение", "Нет отзывов для визуализации!")
                    return None, None

                positive_keywords = []
                negative_keywords = []
                keyword_review_count = Counter()

                for review in all_reviews:
                    if not self.is_running:
                        return None, None
                    sentences = self.analyzer.analyze_review_sentences(review)
                    review_pos_keywords = set()
                    review_neg_keywords = set()
                    for _, _, _, aspects in sentences:
                        for aspect_phrase, aspect_sentiment, _, _ in aspects:
                            if aspect_phrase.lower() in self.analyzer.invalid_phrases:
                                continue
                            if aspect_sentiment == "положительное":
                                positive_keywords.append(aspect_phrase)
                                review_pos_keywords.add(aspect_phrase)
                            elif aspect_sentiment == "отрицательное":
                                negative_keywords.append(aspect_phrase)
                                review_neg_keywords.add(aspect_phrase)
                    for keyword in review_pos_keywords:
                        keyword_review_count[keyword] += 1
                    for keyword in review_neg_keywords:
                        keyword_review_count[keyword] += 1

                pos_keyword_counts = Counter(positive_keywords)
                neg_keyword_counts = Counter(negative_keywords)

                MIN_REVIEW_THRESHOLD = 2
                filtered_pos_keywords = {k: v for k, v in pos_keyword_counts.items() if keyword_review_count[k] >= MIN_REVIEW_THRESHOLD}
                filtered_neg_keywords = {k: v for k, v in neg_keyword_counts.items() if keyword_review_count[k] >= MIN_REVIEW_THRESHOLD}

                return filtered_pos_keywords, filtered_neg_keywords
            except Exception as e:
                self.log(f"Ошибка при вычислении данных для визуализации: {str(e)}", is_error=True)
                messagebox.showerror("Ошибка", f"Не удалось выполнить вычисления: {str(e)}")
                return None, None
            finally:
                self.progress.set(0)

        def show_visualization(filtered_pos_keywords, filtered_neg_keywords):
            if not self.is_running:
                return

            if filtered_pos_keywords is None or filtered_neg_keywords is None:
                return

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

            if filtered_pos_keywords:
                wordcloud_pos = WordCloud(width=800, height=400, background_color='black', colormap='Greens', min_font_size=10, max_words=50).generate_from_frequencies(filtered_pos_keywords)
                ax1.imshow(wordcloud_pos, interpolation='bilinear')
                ax1.set_title("Облако положительных слов", fontsize=16, color='white')
                ax1.axis('off')

            if filtered_neg_keywords:
                wordcloud_neg = WordCloud(width=800, height=400, background_color='black', colormap='Reds', min_font_size=10, max_words=50).generate_from_frequencies(filtered_neg_keywords)
                ax2.imshow(wordcloud_neg, interpolation='bilinear')
                ax2.set_title("Облако отрицательных слов", fontsize=16, color='white')
                ax2.axis('off')

            plt.tight_layout(pad=0)

            top = ctk.CTkToplevel(self.root)
            top.title("Облако слов")
            canvas = FigureCanvasTkAgg(fig, master=top)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)

            self.log("Визуализация облака слов завершена!")

        def run():
            filtered_pos, filtered_neg = compute_data()
            if filtered_pos is not None and filtered_neg is not None and self.is_running:
                self.root.after(0, show_visualization, filtered_pos, filtered_neg)

        thread = threading.Thread(target=run)
        self.threads.append(thread)
        thread.start()

    def log(self, message, is_error=False):
        if not self.is_running:
            return
        try:
            self.log_text.configure(state="normal")
            tag = "ERROR" if is_error else "INFO"
            color = "red" if is_error else "white"
            self.log_text.tag_config(tag, foreground=color)
            self.log_text.insert("end", f"[{time.strftime('%H:%M:%S')}] {tag}: {message}\n", tag)
            self.log_text.configure(state="disabled")
            self.log_text.see("end")
        except Exception:
            pass  # Игнорируем ошибки логирования, если GUI уже закрыт

    def add_link(self, site):
        if not self.is_running:
            return
        link = self.site_links[site].get()
        if link:
            self.log(f"Добавлена ссылка для {SITE_NAMES[site]}: {link}")
        else:
            self.log(f"Поле ссылки для {SITE_NAMES[site]} пустое", is_error=True)

    def setup_logger_redirect(self):
        handler = logging.StreamHandler(self)
        handler.setLevel(logging.INFO)
        handler.addFilter(InfoFilter())
        handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
        logging.getLogger().handlers = [handler]

    def write(self, message):
        if message.strip():
            self.log(message.strip())

    def flush(self):
        pass

    def select_output_dir(self):
        if not self.is_running:
            return
        directory = filedialog.askdirectory()
        if directory:
            self.output_dir.set(directory)

    def clear_log(self):
        if not self.is_running:
            return
        self.log_text.configure(state="normal")
        self.log_text.delete(1.0, ctk.END)
        self.log_text.configure(state="disabled")

    def open_output_dir(self):
        if not self.is_running:
            return
        output_dir = os.path.abspath(self.output_dir.get())
        if os.path.exists(output_dir):
            subprocess.Popen(f'explorer "{output_dir}"' if os.name == 'nt' else ['xdg-open', output_dir])
        else:
            self.log(f"Папка {output_dir} не существует", is_error=True)

    def toggle_parsing(self):
        if not self.is_running:
            return
        if self.running:
            self.running = False
            self.start_btn.configure(text="Старт")
            self.progress.stop()
            self.log("Парсинг остановлен пользователем")
        else:
            self.running = True
            self.start_btn.configure(text="Стоп")
            self.progress.start()
            thread = threading.Thread(target=self.run_parsers)
            self.threads.append(thread)
            thread.start()

    def run_parsers(self):
        parsers = [(name, parser(), link) for site, (name, parser) in {
            "site1": ("Отзовик", Site1Parser),
            "site2": ("iRecommend", Site2Parser),
            "site3": ("Озон", Site3Parser)
        }.items() if (link := self.site_links[site].get())]

        total_steps = len(parsers)

        def run_parser(name, parser, link):
            if not self.running or not self.is_running:
                return
            try:
                self.log(f"Запуск парсера: {name}")
                parser.parse(link, output_dir=self.output_dir.get())
                self.log(f"Парсер {name} завершён")
            except Exception as e:
                self.log(f"Ошибка при работе с {name}: {str(e)}", is_error=True)

        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {executor.submit(run_parser, name, parser, link): name for name, parser, link in parsers}
            for i, future in enumerate(futures):
                future.result()
                self.progress.set((i + 1) / total_steps * 100)

        self.running = False
        self.progress.set(0)
        self.start_btn.configure(text="Старт")
        messagebox.showinfo("Информация", "Парсинг завершён!")

    def select_files_for_analysis_by_site(self):
        if not self.is_running:
            return
        csv_files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if not csv_files:
            self.log("Не выбраны CSV-файлы для анализа!", is_error=True)
            messagebox.showwarning("Предупреждение", "Не выбраны CSV-файлы для анализа!")
            return

        self.csv_files = list(csv_files)
        self.results_by_site = []
        self.result_tree.delete(*self.result_tree.get_children())
        self.log(f"Запуск анализа по сайтам для файлов: {', '.join(csv_files)}")
        thread = threading.Thread(target=self.run_analysis_by_site, args=(csv_files,))
        self.threads.append(thread)
        thread.start()

    def run_analysis_by_site(self, csv_files):
        try:
            self.analyzer = ReviewAnalyzer(use_preprocessing=self.use_preprocessing.get())
            total_files = len(csv_files)
            for i, csv_file in enumerate(csv_files):
                if not self.is_running:
                    return
                try:
                    site_name = self.get_site_name_from_filename(csv_file)
                    if not site_name:
                        self.log(f"Не удалось определить сайт для файла {csv_file}", is_error=True)
                        continue
                    self.log(f"Анализируем файл: {csv_file}")
                    if csv_file in self.analysis_cache:
                        results = self.analysis_cache[csv_file]
                        self.log(f"Использованы кэшированные результаты для {csv_file}")
                    else:
                        results = self.analyzer.analyze_reviews(csv_file)
                        self.analysis_cache[csv_file] = results
                        with open(f"{csv_file}.cache.json", "w", encoding="utf-8") as f:
                            json.dump(results, f, ensure_ascii=False)
                    self.results_by_site.append((site_name, results))
                    self.result_tree.insert("", "end", values=(
                        site_name,
                        results["Плюсы"],
                        results["Минусы"],
                        results["Ключевые слова (положительные)"],
                        results["Ключевые слова (отрицательные)"],
                        results["Общие ключевые слова"],
                        results["Положительные отзывы"],
                        results["Отрицательные отзывы"]
                    ))
                    self.progress.set((i + 1) / total_files * 100)
                except Exception as e:
                    self.log(f"Ошибка при анализе файла {csv_file}: {str(e)}", is_error=True)
                    continue
            self.log("Анализ отзывов по сайтам завершён!")
            self.tabview.set("Анализ слов")
        except Exception as e:
            self.log(f"Ошибка при анализе по сайтам: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось выполнить анализ: {str(e)}")
        finally:
            self.progress.set(0)

    def select_files_for_aggregated_analysis(self):
        if not self.is_running:
            return
        csv_files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if not csv_files:
            self.log("Не выбраны CSV-файлы для анализа!", is_error=True)
            messagebox.showwarning("Предупреждение", "Не выбраны CSV-файлы для анализа!")
            return

        self.csv_files = list(csv_files)
        self.result_tree.delete(*self.result_tree.get_children())
        self.log(f"Запуск агрегированного анализа файлов: {', '.join(csv_files)}")
        thread = threading.Thread(target=self.run_aggregated_analysis, args=(csv_files,))
        self.threads.append(thread)
        thread.start()

    def run_aggregated_analysis(self, csv_files):
        try:
            self.analyzer = ReviewAnalyzer(use_preprocessing=self.use_preprocessing.get())
            cache_key = tuple(sorted(csv_files))
            if cache_key in self.analysis_cache:
                self.aggregated_results = self.analysis_cache[cache_key]
                self.log("Использованы кэшированные результаты для агрегированного анализа")
            else:
                self.aggregated_results = self.analyzer.aggregate_reviews(csv_files)
                self.analysis_cache[cache_key] = self.aggregated_results
                with open("aggregated_analysis.cache.json", "w", encoding="utf-8") as f:
                    json.dump(self.aggregated_results, f, ensure_ascii=False)
            self.result_tree.insert("", "end", values=(
                "Агрегированный анализ",
                self.aggregated_results.get("Плюсы (все сайты)", ""),
                self.aggregated_results.get("Минусы (все сайты)", ""),
                self.aggregated_results.get("Ключевые слова (положительные, все сайты)", ""),
                self.aggregated_results.get("Ключевые слова (отрицательные, все сайты)", ""),
                self.aggregated_results.get("Общие ключевые слова (все сайты)", ""),
                self.aggregated_results.get("Положительные отзывы (все сайты)", ""),
                self.aggregated_results.get("Отрицательные отзывы (все сайты)", "")
            ))
            self.log("Агрегированный анализ отзывов завершён!")
            self.tabview.set("Анализ слов")
        except Exception as e:
            self.log(f"Ошибка при агрегированном анализе: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось выполнить агрегированный анализ: {str(e)}")
        finally:
            self.progress.set(0)

    def select_files_for_detailed_analysis(self):
        if not self.is_running:
            return
        csv_files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if not csv_files:
            self.log("Не выбраны CSV-файлы для анализа!", is_error=True)
            messagebox.showwarning("Предупреждение", "Не выбраны CSV-файлы для анализа!")
            return
        
        self.log("Запуск детального анализа предложений...")
        thread = threading.Thread(target=self.run_detailed_analysis, args=(csv_files,))
        self.threads.append(thread)
        thread.start()

    def run_detailed_analysis(self, csv_files):
        try:
            self.analyzer = ReviewAnalyzer(use_preprocessing=self.use_preprocessing.get())
            self.detailed_tree.delete(*self.detailed_tree.get_children())
            total_files = len(csv_files)
            for i, csv_file in enumerate(csv_files):
                if not self.is_running:
                    return
                try:
                    df = pd.read_csv(csv_file, encoding='utf-8')
                    if 'Текст отзыва' in df.columns:
                        reviews = df['Текст отзыва'].dropna().tolist()
                    elif 'Достоинства' in df.columns and 'Недостатки' in df.columns:
                        reviews = (df['Достоинства'].dropna().tolist() + 
                                   df['Недостатки'].dropna().tolist())
                    else:
                        self.log(f"CSV-файл {csv_file} не содержит подходящих столбцов для анализа!", is_error=True)
                        continue
                    
                    total_reviews = len(reviews)
                    for j, review in enumerate(reviews):
                        if not self.is_running:
                            return
                        sentences_analysis = self.analyzer.analyze_review_sentences(review)
                        for sentence, sentiment, score, aspects in sentences_analysis:
                            aspects_str = "; ".join([f"{aspect}: {sent} ({s:.2f})" for aspect, sent, s, _ in aspects])
                            self.detailed_tree.insert("", "end", values=(
                                sentence,
                                sentiment,
                                f"{score:.2f}",
                                aspects_str
                            ))
                        self.progress_detailed.set((j + 1) / total_reviews * 100)
                    self.progress_detailed.set(0)
                    self.log(f"Детальный анализ предложений для {csv_file} завершён!")
                except Exception as e:
                    self.log(f"Ошибка при детальном анализе файла {csv_file}: {str(e)}", is_error=True)
                    messagebox.showerror("Ошибка", f"Не удалось обработать файл {csv_file}: {str(e)}")
                    continue
                self.progress.set((i + 1) / total_files * 100)
            self.tabview.set("Детальный анализ предложений")
        except Exception as e:
            self.log(f"Ошибка при детальном анализе: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось выполнить детальный анализ: {str(e)}")
        finally:
            self.progress.set(0)
            self.progress_detailed.set(0)

    def generate_report_by_site(self):
        if not self.is_running:
            return
        if not hasattr(self, 'results_by_site') or not self.results_by_site:
            self.log("Сначала выполните анализ по сайтам на вкладке 'Анализ слов'!", is_error=True)
            messagebox.showwarning("Предупреждение", "Сначала выполните анализ по сайтам!")
            return

        self.log("Генерация отчёта по сайтам...")
        thread = threading.Thread(target=self.run_generate_report_by_site)
        self.threads.append(thread)
        thread.start()

    def run_generate_report_by_site(self):
        try:
            self.progress_report.set(0)
            self.report_textbox.delete("0.0", "end")
            product_name = self.product_name.get().strip() or "продукт"
            full_report = ""
            total_sites = len(self.results_by_site)
            for i, (site_name, results) in enumerate(self.results_by_site):
                if not self.is_running:
                    return
                full_report += f"Отчёт для {site_name}:\n\n"
                report = self.analyzer.generate_detailed_report(results, product_name=product_name)
                full_report += report + "\n" + "="*50 + "\n\n"
                self.progress_report.set((i + 1) / total_sites * 100)
            self.report_textbox.insert("0.0", full_report)
            self.log("Текстовый отчёт по сайтам сгенерирован!")
            self.tabview.set("Текстовый отчёт")
        except Exception as e:
            self.log(f"Ошибка при генерации отчёта по сайтам: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось сгенерировать отчёт: {str(e)}")
        finally:
            self.progress_report.set(0)

    def generate_aggregated_report(self):
        if not self.is_running:
            return
        if not hasattr(self, 'aggregated_results'):
            self.log("Сначала выполните агрегированный анализ на вкладке 'Анализ слов'!", is_error=True)
            messagebox.showwarning("Предупреждение", "Сначала выполните агрегированный анализ!")
            return

        self.log("Генерация агрегированного отчёта...")
        thread = threading.Thread(target=self.run_generate_aggregated_report)
        self.threads.append(thread)
        thread.start()

    def run_generate_aggregated_report(self):
        try:
            self.progress_report.set(0)
            self.report_textbox.delete("0.0", "end")
            product_name = self.product_name.get().strip() or "продукт"
            report = self.analyzer.generate_detailed_report(self.aggregated_results, product_name=product_name)
            self.report_textbox.insert("0.0", report)
            self.progress_report.set(100)
            self.log("Агрегированный текстовый отчёт сгенерирован!")
            self.tabview.set("Текстовый отчёт")
        except Exception as e:
            self.log(f"Ошибка при генерации агрегированного отчёта: {str(e)}", is_error=True)
            messagebox.showerror("Ошибка", f"Не удалось сгенерировать отчёт: {str(e)}")
        finally:
            self.progress_report.set(0)

    def get_site_name_from_filename(self, filename):
        base_name = os.path.basename(filename).lower()
        if "otzovik" in base_name:
            return SITE_NAMES["site1"]
        elif "irecommend" in base_name:
            return SITE_NAMES["site2"]
        elif "ozon" in base_name:
            return SITE_NAMES["site3"]
        return None

    def bind_shortcuts(self):
        def handle_ctrl_key(event):
            if event.state & 0x4:
                if event.keycode == 67:  # Ctrl + C
                    event.widget.event_generate("<<Copy>>")
                    return "break"
                elif event.keycode == 86:  # Ctrl + V
                    event.widget.event_generate("<<Paste>>")
                    return "break"
                elif event.keycode == 88:  # Ctrl + X
                    event.widget.event_generate("<<Cut>>")
                    return "break"
                elif event.keycode == 65:  # Ctrl + A
                    self.select_all(event.widget)
                    return "break"
                elif event.keycode == 90:  # Ctrl + Z
                    if isinstance(event.widget, ctk.CTkEntry):
                        event.widget.delete(0, ctk.END)
                    return "break"

        def bind_to_entry(widget):
            if isinstance(widget, (ctk.CTkEntry, ctk.CTkTextbox)):
                widget.bind("<Control-Key>", handle_ctrl_key)
            for child in widget.winfo_children():
                bind_to_entry(child)

        bind_to_entry(self.root)

    def select_all(self, widget):
        if isinstance(widget, ctk.CTkEntry):
            widget.select_range(0, 'end')
            widget.icursor('end')
        elif isinstance(widget, ctk.CTkTextbox):
            widget.tag_add("sel", "1.0", "end")
    
    def load_config(self):
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                self.output_dir.set(config.get("output_dir", DATA_DIR))
                for site, link in config.get("site_links", {}).items():
                    if site in self.site_links:
                        self.site_links[site].set(link)
        except FileNotFoundError:
            pass

    def save_config(self):
        config = {
            "output_dir": self.output_dir.get(),
            "site_links": {site: link.get() for site, link in self.site_links.items()}
        }
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f)

    def on_closing(self):
        self.is_running = False  # Останавливаем все операции
        self.running = False  # Останавливаем парсинг

        # Ждём завершения всех потоков
        for thread in self.threads:
            if thread.is_alive():
                thread.join(timeout=2.0)  # Даём потокам 2 секунды на завершение

        self.save_config()

        # Завершаем цикл событий Tkinter
        self.root.quit()

if __name__ == "__main__":
    root = ctk.CTk()
    app = ParserApp(root)
    root.mainloop()